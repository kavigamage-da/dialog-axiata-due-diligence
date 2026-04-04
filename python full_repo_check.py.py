# full_repo_check.py
import os
import re
import openpyxl

# Folder to scan (current folder)
base_folder = "."

# List all files to scan
file_types = (".py", ".md", ".R", ".xlsx")
all_files = []

for root, dirs, files in os.walk(base_folder):
    for file in files:
        if file.endswith(file_types):
            all_files.append(os.path.join(root, file))

# Reference numeric values to verify
reference_values = [
    66276, 38.7, 24.5, -18.8, 171.2, 1.38,
    233, 570.6, 220.9, 7.26, 38.72, 1.38,  # your table example
    4624.9, 2630.4, 1115.8, 3.1, 42.42, 70.03, 1.31, 225,
    130561.8, 24367.2, 12678.6, 14.94, 52.03, 117.09, 23.12, 35.33,
    4245.7, 3577.1, 1329.2, 9.75, 37.16, 138.99, 15.28, 12.18,
    19624.3, 9326.9, 3895.8, 14.77, 41.77, 50.11, 18.31, 14.27
]

def check_text_file(file_path):
    """Scan text-based files for numeric values."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    found = {str(v): False for v in reference_values}
    for v in reference_values:
        pattern = re.escape(str(v))
        if re.search(pattern, text):
            found[str(v)] = True
    return found

def check_excel_file(file_path):
    """Scan all sheets in Excel for numeric values."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    found = {str(v): False for v in reference_values}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for v in reference_values:
                    if str(v) in str(cell):
                        found[str(v)] = True
    return found

# Main check
for file_path in all_files:
    ext = os.path.splitext(file_path)[1].lower()
    print(f"\nFile: {file_path}")
    try:
        if ext == ".xlsx":
            result = check_excel_file(file_path)
        else:
            result = check_text_file(file_path)
        for v, f in result.items():
            print(f"  {v}: {'FOUND ✅' if f else 'MISSING ❌'}")
    except Exception as e:
        print(f"  ⚠ Error checking file: {e}")