# check_excel_numbers.py
import openpyxl

# Files to check
files = [
    "data/competitor_benchmarking.xlsx",
    "model/Dialog_Axiata_Financial_Model.xlsx"
]

# Numbers to verify
values = ["66276", 38.7, 24.5, -18.8, 171.2, 1.38]

for file in files:
    wb = openpyxl.load_workbook(file, data_only=True)
    print(f"\nChecking {file}:")
    found = {str(v): False for v in values}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for v in values:
                    if str(v) in str(cell):
                        found[str(v)] = True
    
    # Report
    for v, f in found.items():
        print(f"  {v}: {'FOUND ✅' if f else 'MISSING ❌'}")