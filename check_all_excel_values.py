# check_all_excel_values.py
import openpyxl

# List of Excel files to check
files = [
    r"D:\Dialog axiata financial Analyst\competitor_benchmarking.xlsx",
    r"D:\Dialog axiata financial Analyst\model\Dialog_Axiata_Financial_Model.xlsx"
]
# Numbers/values to verify
values = ["66276", 38.7, 24.5, -18.8, 171.2, 1.38]

for filepath in files:
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"Error opening {filepath}: {e}")
        continue

    print(f"\nChecking {filepath}:")
    found = {str(v): False for v in values}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                for v in values:
                    if str(v) in str(cell):
                        found[str(v)] = True

    # Report
    for v, f in found.items():
        print(f"  {v}: {'FOUND ✅' if f else 'MISSING ❌'}")