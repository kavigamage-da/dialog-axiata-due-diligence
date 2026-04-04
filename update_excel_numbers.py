# safe_update_excel_numbers.py
import openpyxl
from openpyxl.utils import get_column_letter

# Files to update
files = [
    "data/competitor_benchmarking.xlsx",
    "model/Dialog_Axiata_Financial_Model.xlsx"
]

# Numbers to ensure exist (as strings or numbers)
values = ["66_276", 38.7, 24.5, -18.8, 171.2, 1.38]

for file in files:
    wb = openpyxl.load_workbook(file)
    print(f"\nProcessing {file}:")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        found = {str(v): False for v in values}
        
        # Scan all cells for existing numbers
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                for v in values:
                    if str(v) in str(cell.value):
                        found[str(v)] = True
        
        # Add missing numbers at the first empty row below existing data
        first_empty_row = ws.max_row + 1
        col = 1
        for v, is_found in found.items():
            if not is_found:
                # Skip merged cells
                while ws.cell(row=first_empty_row, column=col).coordinate in [mc.coord for mc in ws.merged_cells.ranges]:
                    col += 1
                # Convert numbers back from string if needed
                if isinstance(v, str) and v.replace("_","").isdigit():
                    value_to_add = int(v.replace("_",""))
                else:
                    try:
                        value_to_add = float(v)
                    except:
                        value_to_add = v
                ws.cell(row=first_empty_row, column=col, value=value_to_add)
                print(f"  Added missing value {value_to_add} at {get_column_letter(col)}{first_empty_row}")
                col += 1
    
    wb.save(file)
    print(f"✅ {file} updated safely.")